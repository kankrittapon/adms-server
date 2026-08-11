"""
Excel Human Master Import Utility
PromptID: ADMS-Data-ExcelImport-002

Parses the approved Excel Human Master workbook and atomically reconciles
personnel records into human_employees and human_employee_sources.

Default mode is --dry-run. Explicit --apply flag required for PostgreSQL transaction commit.
This script DOES NOT perform Human <-> Device mapping or ZKTeco network calls.
"""

import os
import sys
import argparse
import hashlib
import re
from datetime import datetime
import openpyxl
import psycopg2
from psycopg2.extras import RealDictCursor

# Ensure stdout/stderr handles UTF-8 strings cleanly on Windows
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')
if sys.stderr.encoding != 'utf-8':
    sys.stderr.reconfigure(encoding='utf-8')

# Import DB configuration from app context
try:
    from app.config import Config
except ImportError:
    class Config:
        DB_HOST = os.getenv("DB_HOST", "localhost")
        DB_PORT = int(os.getenv("DB_PORT", "5432"))
        DB_NAME = os.getenv("DB_NAME", "adms")
        DB_USER = os.getenv("DB_USER", "postgres")
        DB_PASSWORD = os.getenv("DB_PASSWORD", "postgres")

DEFAULT_WORKBOOK_PATH = os.path.join("excel", "files", "รายละเอียด กพ.พัน.สอล.ฯ ก.พ.69.xlsx")
DEFAULT_SHEET_NAME = "ยอด ม.ค.69"
SOURCE_SYSTEM = "EXCEL_HUMAN_MASTER"

# Known rank prefixes to strip from display name
RANK_PREFIXES = [
    "ว่าที่ น.ต.", "ว่าที่ ร.ต.",
    "น.ท.", "น.ต.", "ร.อ.", "ร.ท.", "ร.ต.",
    "พ.จ.อ.", "พ.จ.ท.", "พ.จ.ต.",
    "จ.อ.", "จ.ท.", "จ.ต.",
    "พลฯ", "พลทหาร"
]

def normalize_text(text: str) -> str:
    """Trims whitespace and normalizes internal spaces."""
    if not text:
        return ""
    text = str(text).strip()
    return re.sub(r'\s+', ' ', text)

def extract_rank_and_name(raw_name_cell: str) -> tuple[str, str]:
    """Extracts rank prefix and clean display name from name cell."""
    clean_raw = normalize_text(raw_name_cell)
    if not clean_raw:
        return "", ""
    
    for prefix in RANK_PREFIXES:
        if clean_raw.startswith(prefix):
            name_part = clean_raw[len(prefix):].strip()
            return prefix, name_part
            
    # Fallback if no known prefix matched
    parts = clean_raw.split(None, 1)
    if len(parts) == 2 and len(parts[0]) <= 8:
        return parts[0], parts[1]
        
    return "", clean_raw

def compute_source_hash(rank: str, display_name: str, branch: str, category: str, notes: str) -> str:
    """Computes deterministic SHA256 content hash for a normalized personnel row."""
    normalized_string = f"{normalize_text(rank)}|{normalize_text(display_name)}|{normalize_text(branch)}|{normalize_text(category)}|{normalize_text(notes)}"
    return hashlib.sha256(normalized_string.encode('utf-8')).hexdigest()

def parse_workbook(file_path: str, sheet_name: str = DEFAULT_SHEET_NAME) -> list[dict]:
    """Parses Excel workbook and returns normalized personnel records."""
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Source workbook not found at: {file_path}")
        
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook. Available: {wb.sheetnames}")
        
    sheet = wb[sheet_name]
    records = []
    
    current_category = ""
    category_idx = 0
    
    # Category detection markers in source workbook
    category_headers = {
        "นายทหาร": 1,
        "พันจ่า": 2,
        "จ่า": 3,
        "พลทหาร": 4
    }
    
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        cell_1 = normalize_text(row[0]) if len(row) > 0 and row[0] is not None else ""
        cell_2 = normalize_text(row[1]) if len(row) > 1 and row[1] is not None else ""
        cell_3 = normalize_text(row[2]) if len(row) > 2 and row[2] is not None else ""
        cell_4 = normalize_text(row[3]) if len(row) > 3 and row[3] is not None else ""
        
        # Check if row is a category header
        for cat_name, cat_id in category_headers.items():
            if cat_name in cell_1 or cat_name in cell_2:
                current_category = cat_name
                category_idx = cat_id
                break
                
        # Skip header rows or summary rows without numerical 'ที่'
        if not cell_1.isdigit():
            continue
            
        if not current_category:
            continue
            
        seq_num = int(cell_1)
        rank, display_name = extract_rank_and_name(cell_2)
        branch = cell_3
        notes = cell_4
        
        source_record_key = f"EXCEL_FEB69_CAT_{category_idx}_ROW_{row_idx:03d}"
        source_hash = compute_source_hash(rank, display_name, branch, current_category, notes)
        
        records.append({
            "source_row": row_idx,
            "seq_num": seq_num,
            "rank": rank,
            "display_name": display_name,
            "raw_name": cell_2,
            "branch": branch,
            "category": current_category,
            "category_idx": category_idx,
            "notes": notes,
            "source_record_key": source_record_key,
            "source_hash": source_hash,
            "source_file": os.path.relpath(file_path).replace("\\", "/"),
            "source_sheet": sheet_name
        })
        
    return records

def get_db_connection():
    """Establishes database connection using application configuration or environment defaults."""
    try:
        cfg = Config.from_env()
        return psycopg2.connect(
            host=cfg.db_host,
            port=cfg.db_port,
            dbname=cfg.db_name,
            user=cfg.db_user,
            password=cfg.db_password
        )
    except Exception:
        host = os.getenv("DB_HOST", "localhost")
        port = int(os.getenv("DB_PORT", "5432"))
        dbname = os.getenv("DB_NAME", "adms")
        user = os.getenv("DB_USER", "postgres")
        password = os.getenv("DB_PASSWORD", "postgres")
        return psycopg2.connect(host=host, port=port, dbname=dbname, user=user, password=password)

def reconcile_import(records: list[dict], apply: bool = False) -> dict:
    """Performs dry-run or atomic database import reconciliation."""
    summary = {
        "total_parsed": len(records),
        "categories": {},
        "valid": 0,
        "invalid": 0,
        "new": 0,
        "unchanged": 0,
        "changed": 0,
        "ambiguous": 0,
        "applied": apply
    }
    
    for r in records:
        cat = r["category"]
        summary["categories"][cat] = summary["categories"].get(cat, 0) + 1
        if r["display_name"]:
            summary["valid"] += 1
        else:
            summary["invalid"] += 1
            
    if summary["invalid"] > 0:
        raise ValueError(f"Import aborted: {summary['invalid']} records missing display_name")
        
    if not apply:
        # Dry-Run query against DB to calculate NEW vs UNCHANGED vs CHANGED
        conn = None
        try:
            conn = get_db_connection()
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("SELECT source_record_key, source_hash FROM human_employee_sources WHERE source_system = %s", (SOURCE_SYSTEM,))
                existing = {row["source_record_key"]: row["source_hash"] for row in cur.fetchall()}
                
                for r in records:
                    key = r["source_record_key"]
                    h = r["source_hash"]
                    if key not in existing:
                        summary["new"] += 1
                    elif existing[key] == h:
                        summary["unchanged"] += 1
                    else:
                        summary["changed"] += 1
        except Exception as e:
            print(f"[DRY-RUN] DB query note: Could not connect to DB ({e}). Assuming clean DB (all NEW).")
            summary["new"] = summary["total_parsed"]
        finally:
            if conn:
                conn.close()
        return summary
        
    # Apply Mode: Atomic PostgreSQL Transaction
    conn = get_db_connection()
    
    try:
        with conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                for r in records:
                    key = r["source_record_key"]
                    h = r["source_hash"]
                    
                    cur.execute("""
                        SELECT s.employee_id, s.source_hash 
                        FROM human_employee_sources s 
                        WHERE s.source_system = %s AND s.source_record_key = %s
                    """, (SOURCE_SYSTEM, key))
                    existing = cur.fetchone()
                    
                    if not existing:
                        # Case A: NEW record -> INSERT human_employees + INSERT human_employee_sources
                        cur.execute("""
                            INSERT INTO human_employees (display_name, rank, branch, category, notes)
                            VALUES (%s, %s, %s, %s, %s)
                            RETURNING employee_id
                        """, (r["display_name"], r["rank"], r["branch"], r["category"], r["notes"]))
                        emp_id = cur.fetchone()["employee_id"]
                        
                        cur.execute("""
                            INSERT INTO human_employee_sources (
                                employee_id, source_system, source_file, source_sheet, 
                                source_row, source_record_key, source_hash
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s)
                        """, (
                            emp_id, SOURCE_SYSTEM, r["source_file"], r["source_sheet"],
                            r["source_row"], key, h
                        ))
                        summary["new"] += 1
                        
                    elif existing["source_hash"] == h:
                        # Case B: UNCHANGED -> SKIP
                        summary["unchanged"] += 1
                        
                    else:
                        # Case C: CHANGED -> UPDATE human_employees + UPDATE human_employee_sources
                        emp_id = existing["employee_id"]
                        cur.execute("""
                            UPDATE human_employees 
                            SET display_name = %s, rank = %s, branch = %s, category = %s, notes = %s, updated_at = now()
                            WHERE employee_id = %s
                        """, (r["display_name"], r["rank"], r["branch"], r["category"], r["notes"], emp_id))
                        
                        cur.execute("""
                            UPDATE human_employee_sources
                            SET source_hash = %s, source_file = %s, source_sheet = %s, source_row = %s, updated_at = now()
                            WHERE source_system = %s AND source_record_key = %s
                        """, (h, r["source_file"], r["source_sheet"], r["source_row"], SOURCE_SYSTEM, key))
                        summary["changed"] += 1
                        
    finally:
        conn.close()
        
    return summary

def main():
    parser = argparse.ArgumentParser(description="Excel Human Master Import Utility")
    parser.add_argument("--file", default=DEFAULT_WORKBOOK_PATH, help="Path to source Excel workbook")
    parser.add_argument("--sheet", default=DEFAULT_SHEET_NAME, help="Source sheet name")
    parser.add_argument("--apply", action="store_true", help="Explicitly commit database changes (default is dry-run)")
    parser.add_argument("--dry-run", action="store_true", help="Perform dry-run validation only (default)")
    
    args = parser.parse_args()
    
    is_apply = args.apply and not args.dry_run
    mode_str = "APPLY (COMMIT TO DATABASE)" if is_apply else "DRY-RUN (READ ONLY)"
    
    print(f"=== Excel Human Master Import Utility ===")
    print(f"Mode: {mode_str}")
    print(f"Workbook: {args.file}")
    print(f"Sheet: {args.sheet}\n")
    
    records = parse_workbook(args.file, args.sheet)
    print(f"Parsed {len(records)} personnel records cleanly.\n")
    
    # Category summary
    summary = reconcile_import(records, apply=is_apply)
    print("--- CATEGORY BREAKDOWN ---")
    for cat, count in summary["categories"].items():
        print(f"  - {cat:12s}: {count:3d} personnel")
    print(f"  TOTAL       : {summary['total_parsed']:3d} personnel\n")
    
    print("--- RECONCILIATION SUMMARY ---")
    print(f"  Valid records      : {summary['valid']}")
    print(f"  Invalid records    : {summary['invalid']}")
    print(f"  NEW records        : {summary['new']}")
    print(f"  UNCHANGED records  : {summary['unchanged']}")
    print(f"  CHANGED records    : {summary['changed']}")
    print(f"  AMBIGUOUS records  : {summary['ambiguous']}")
    print(f"  Applied to DB      : {summary['applied']}\n")
    
    # Print sample records
    print("--- SAMPLE RECORD PREVIEW (First 2 per category) ---")
    seen_cats = {}
    for r in records:
        cat = r["category"]
        seen_cats[cat] = seen_cats.get(cat, 0) + 1
        if seen_cats[cat] <= 2:
            print(f"  [{r['source_record_key']}] Cat: {r['category']:8s} | Rank: {r['rank']:10s} | Name: {r['display_name']:20s} | Branch: {r['branch']:6s} | Notes: {r['notes']}")
            
    if not is_apply:
        print("\n[INFO] Dry-run complete. No database changes were made. Use --apply to execute PostgreSQL transaction.")
    else:
        print("\n[SUCCESS] Atomic database import completed successfully!")

if __name__ == "__main__":
    main()
